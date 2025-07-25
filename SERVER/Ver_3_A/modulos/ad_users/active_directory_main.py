#!/usr/bin/env python3
"""
Script para importar dados de usuários do Active Directory
de uma planilha Excel para banco de dados PostgreSQL
"""

import openpyxl
import logging
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError
import sys
from typing import Optional, Dict, Any
import os

# Importar os modelos
from active_directory_models.ad_user_database import Base, ADUser

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ad_import.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class ADImporter:
    """Classe para importar dados do AD de Excel para PostgreSQL"""

    def __init__(self, db_url: str):
        """
        Inicializa o importador

        Args:
            db_url: URL de conexão do PostgreSQL
                   Formato: postgresql://usuario:senha@host:porta/banco
        """
        self.engine = create_engine(db_url, echo=False)
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.session = Session()
        logger.info("Conexão com banco de dados estabelecida")

    def parse_datetime(self, date_value) -> Optional[datetime]:
        """Converte valores de data/hora para o formato correto"""
        if date_value is None or date_value == '':
            return None

        # Se já for datetime, retorna
        if isinstance(date_value, datetime):
            return date_value

        # Tentar converter string para datetime
        if isinstance(date_value, str):
            try:
                # Tenta diferentes formatos comuns
                for fmt in ['%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S',
                            '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        return datetime.strptime(date_value, fmt)
                    except ValueError:
                        continue
            except Exception as e:
                logger.warning(f"Erro ao converter data '{date_value}': {e}")
                return None

        return None

    def parse_boolean(self, bool_value) -> Optional[bool]:
        """Converte valores booleanos"""
        if bool_value is None or bool_value == '':
            return None

        # Se já for booleano
        if isinstance(bool_value, bool):
            return bool_value

        # Converter string
        if isinstance(bool_value, str):
            return bool_value.lower() in ['true', 'yes', '1', 'enabled', 'sim', 'verdadeiro']

        # Converter número
        if isinstance(bool_value, (int, float)):
            return bool(bool_value)

        return None

    def import_from_excel(self, excel_path: str, sheet_name: str = None):
        """
        Importa dados de uma planilha Excel usando openpyxl

        Args:
            excel_path: Caminho para o arquivo Excel
            sheet_name: Nome da planilha (opcional)
        """
        try:
            logger.info(f"Lendo arquivo Excel: {excel_path}")

            # Abrir o arquivo Excel
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

            # Selecionar a planilha
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(
                        f"Planilha '{sheet_name}' não encontrada. Planilhas disponíveis: {workbook.sheetnames}")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            # Obter os cabeçalhos (primeira linha)
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append('')

            logger.info(f"Colunas encontradas: {headers}")

            # Mapear colunas para índices
            column_mapping = {
                'displayname': 'DisplayName',
                'samaccountname': 'SamAccountName',
                'givenname': 'GivenName',
                'surname': 'Surname',
                'emailaddress': 'EmailAddress',
                'enabled': 'Enabled',
                'lastlogondate': 'LastLogonDate',
                'distinguishedname': 'distinguishedName'
            }

            # Criar mapa de índices das colunas
            column_indices = {}
            for idx, header in enumerate(headers):
                header_lower = header.lower().replace(' ', '').replace('_', '')
                if header_lower in column_mapping:
                    column_indices[column_mapping[header_lower]] = idx

            # Verificar se a coluna obrigatória existe
            if 'SamAccountName' not in column_indices:
                raise ValueError("Coluna obrigatória 'SamAccountName' não encontrada na planilha")

            # Contar total de linhas (pulando o cabeçalho)
            total_rows = worksheet.max_row - 1
            logger.info(f"Total de registros encontrados: {total_rows}")

            # Processar cada linha (começando da linha 2)
            imported = 0
            updated = 0
            errors = 0

            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Criar dicionário com os dados da linha
                    row_data = {}
                    for col_name, col_idx in column_indices.items():
                        if col_idx < len(row):
                            row_data[col_name] = row[col_idx]
                        else:
                            row_data[col_name] = None

                    # Verificar se o usuário já existe
                    sam_account = str(row_data.get('SamAccountName', '') or '').strip()
                    if not sam_account:
                        logger.warning(f"Linha {row_num}: SamAccountName vazio, pulando...")
                        continue

                    existing_user = self.session.query(ADUser).filter_by(
                        sam_account_name=sam_account
                    ).first()

                    if existing_user:
                        # Atualizar usuário existente
                        existing_user.display_name = str(row_data.get('DisplayName', '') or '').strip() or None
                        existing_user.given_name = str(row_data.get('GivenName', '') or '').strip() or None
                        existing_user.surname = str(row_data.get('Surname', '') or '').strip() or None
                        existing_user.email_address = str(row_data.get('EmailAddress', '') or '').strip() or None
                        existing_user.enabled = self.parse_boolean(row_data.get('Enabled'))
                        existing_user.last_logon_date = self.parse_datetime(row_data.get('LastLogonDate'))
                        existing_user.distinguished_name = str(
                            row_data.get('distinguishedName', '') or '').strip() or None
                        updated += 1
                        logger.debug(f"Usuário atualizado: {sam_account}")
                    else:
                        # Criar novo usuário
                        new_user = ADUser(
                            sam_account_name=sam_account,
                            display_name=str(row_data.get('DisplayName', '') or '').strip() or None,
                            given_name=str(row_data.get('GivenName', '') or '').strip() or None,
                            surname=str(row_data.get('Surname', '') or '').strip() or None,
                            email_address=str(row_data.get('EmailAddress', '') or '').strip() or None,
                            enabled=self.parse_boolean(row_data.get('Enabled')),
                            last_logon_date=self.parse_datetime(row_data.get('LastLogonDate')),
                            distinguished_name=str(row_data.get('distinguishedName', '') or '').strip() or None
                        )
                        self.session.add(new_user)
                        imported += 1
                        logger.debug(f"Novo usuário adicionado: {sam_account}")

                    # Commit a cada 100 registros
                    if (imported + updated) % 100 == 0:
                        self.session.commit()
                        logger.info(f"Progresso: {imported} importados, {updated} atualizados")

                except Exception as e:
                    errors += 1
                    logger.error(f"Erro na linha {row_num}: {str(e)}")
                    self.session.rollback()
                    continue

            # Commit final
            self.session.commit()

            # Fechar o workbook
            workbook.close()

            logger.info("=" * 50)
            logger.info(f"Importação concluída!")
            logger.info(f"Novos registros: {imported}")
            logger.info(f"Registros atualizados: {updated}")
            logger.info(f"Erros: {errors}")
            logger.info("=" * 50)

            return {
                'imported': imported,
                'updated': updated,
                'errors': errors,
                'total': total_rows
            }

        except Exception as e:
            logger.error(f"Erro durante importação: {str(e)}")
            self.session.rollback()
            raise
        finally:
            self.session.close()


def main():
    """Função principal"""
    # Configurações - AJUSTE CONFORME NECESSÁRIO
    DB_USER = os.getenv('DB_USER', 'isac')
    DB_PASSWORD = os.getenv('DB_PASSWORD', 'kwa44fgjc8suf91kjsacaz')
    DB_HOST = os.getenv('DB_HOST', 'localhost')
    DB_PORT = os.getenv('DB_PORT', '5432')
    DB_NAME = os.getenv('DB_NAME', 'DB_USERS')

    # Caminho do arquivo Excel
    EXCEL_FILE = os.getenv('EXCEL_FILE', 'ListaDeUsuariosDoDominio.xlsx')
    SHEET_NAME = os.getenv('SHEET_NAME', None)  # None para usar a primeira planilha

    # Montar URL de conexão
    db_url = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

    try:
        # Verificar se o arquivo existe
        if not os.path.exists(EXCEL_FILE):
            logger.error(f"Arquivo não encontrado: {EXCEL_FILE}")
            sys.exit(1)

        # Criar importador e executar
        importer = ADImporter(db_url)
        results = importer.import_from_excel(EXCEL_FILE, SHEET_NAME)

        # Exibir resumo
        print("\nResumo da importação:")
        print(f"- Total de linhas no Excel: {results['total']}")
        print(f"- Novos registros importados: {results['imported']}")
        print(f"- Registros atualizados: {results['updated']}")
        print(f"- Erros encontrados: {results['errors']}")

    except Exception as e:
        logger.error(f"Erro fatal: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()